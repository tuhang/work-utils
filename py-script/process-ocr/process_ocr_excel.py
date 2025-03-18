import os
import requests
import pandas as pd
from tqdm import tqdm
import time
import io
from PIL import Image
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
import concurrent.futures
import threading
import glob

# 创建一个线程锁，用于安全更新DataFrame
df_lock = threading.Lock()

# 创建图片保存目录
if not os.path.exists('images'):
    os.makedirs('images')

# 查找文件夹下所有尾部包含"-个人"的Excel文件
excel_files = []
for file in glob.glob("*.xlsx"):
    if file.endswith("-个人.xlsx"):
        excel_files.append(file)

if not excel_files:
    print("错误：未找到符合条件的Excel文件（文件名尾部包含'-个人'）")
    exit(1)

print(f"找到以下符合条件的文件：{excel_files}")

# 处理每个Excel文件
for excel_file in excel_files:
    print(f"\n开始处理文件: {excel_file}")
    
    # 读取Excel文件
    df = pd.read_excel(excel_file)

    # 确保dz_pic_url_id列存在
    if 'dz_pic_url_id' not in df.columns:
        print(f"警告：文件 {excel_file} 中没有找到'dz_pic_url_id'列，跳过此文件")
        continue

    # 确保subject_id列使用字符串类型
    if 'subject_id' in df.columns:
        df['subject_id'] = df['subject_id'].astype(str)

    # 获取dz_pic_url_id列的索引位置
    dz_pic_url_id_idx = df.columns.get_loc('dz_pic_url_id')
    
    # 添加"用户作答"列在dz_pic_url_id右侧
    column_list = list(df.columns)
    column_list.insert(dz_pic_url_id_idx + 1, '用户作答')
    column_list.insert(dz_pic_url_id_idx + 2, '复核状态')
    column_list.insert(dz_pic_url_id_idx + 3, '复核备注')
    df = df.reindex(columns=column_list)
    
    # 设置默认值
    df['复核状态'] = 'SKIP'
    df['复核备注'] = ''
    
    # 添加picFile列（内部使用）
    if 'picFile' not in df.columns:
        df['picFile'] = ''

    # API请求URL和头信息
    url = 'todoUrl'
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/58.0.3029.110 Safari/537.3',
        'Content-Type': 'application/json',
        'Accept': '*/*',
        'Host': 'todoUrlHost',
        'Connection': 'keep-alive'
    }

    # 检查现有图片缓存
    print("检查现有图片缓存...")
    existing_images = {}
    for img_path in glob.glob("images/*"):
        filename = os.path.basename(img_path)
        # 从文件名中提取dz_pic_url_id
        parts = filename.split('_')
        if len(parts) >= 2:
            dz_pic_url_id = parts[-1].split('.')[0]  # 获取ID部分，去掉扩展名
            existing_images[dz_pic_url_id] = img_path

    print(f"找到 {len(existing_images)} 个缓存图片")

    # 处理单个图片的函数
    def process_image(args):
        index, row = args
        dz_pic_url_id = row['dz_pic_url_id']
        answer_id = row['answerId'] if 'answerId' in df.columns else f"item_{index}"
        
        # 跳过空值
        if pd.isna(dz_pic_url_id) or dz_pic_url_id == '':
            return None
        
        # 检查缓存中是否已存在该图片
        for key, path in existing_images.items():
            if dz_pic_url_id in key:
                print(f"使用缓存图片: {path}")
                # 安全地更新DataFrame
                with df_lock:
                    df.at[index, 'picFile'] = path
                    df.at[index, '用户作答'] = path
                return path
        
        # 准备请求数据
        data = {
            "dzPicUrlId": dz_pic_url_id
        }
        
        try:
            # 发送请求获取图片URL
            response = requests.post(url, headers=headers, json=data)
            response.raise_for_status()  # 检查请求是否成功
            
            # 解析响应
            result = response.json()
            
            # 修改判断条件：code为0也是成功
            if ('code' in result and (result['code'] == 200 or result['code'] == 0)) and 'data' in result:
                # 获取图片URL
                image_url = result['data']
                
                # 如果返回的是对象而不是直接的URL，需要提取URL
                if isinstance(image_url, dict) and 'url' in image_url:
                    image_url = image_url['url']
                
                # 下载图片
                img_response = requests.get(image_url)
                img_response.raise_for_status()
                
                # 保存图片
                file_extension = image_url.split('.')[-1].split('?')[0]  # 获取文件扩展名
                if not file_extension or len(file_extension) > 5:
                    file_extension = 'jpg'  # 默认使用jpg
                    
                image_filename = f"images/{answer_id}_{dz_pic_url_id}.{file_extension}"
                with open(image_filename, 'wb') as f:
                    f.write(img_response.content)
                
                # 安全地更新DataFrame
                with df_lock:
                    df.at[index, 'picFile'] = image_filename
                    df.at[index, '用户作答'] = image_filename
                
                return image_filename
            else:
                print(f"获取图片URL失败: {dz_pic_url_id}, 响应: {result}")
                return None
        except Exception as e:
            print(f"处理图片时出错 {dz_pic_url_id}: {str(e)}")
            return None

    # 使用线程池处理图片
    print("开始多线程下载图片...")
    with concurrent.futures.ThreadPoolExecutor(max_workers=20) as executor:
        # 创建任务列表
        tasks = [(index, row) for index, row in df.iterrows()]
        
        # 提交任务并显示进度条
        results = list(tqdm(
            executor.map(process_image, tasks), 
            total=len(tasks),
            desc="下载图片"
        ))

    # 生成输出文件名
    output_base = os.path.splitext(excel_file)[0]
    output_file = f'{output_base}_with_images.xlsx'
    
    # 保存更新后的Excel文件（先不包含图片）
    df.to_excel(output_file, index=False)
    print(f"基本数据已保存到 {output_file}")

    # 现在添加图片到Excel
    print("正在将图片嵌入到Excel中...")

    try:
        # 加载工作簿
        wb = load_workbook(output_file)
        ws = wb.active

        # 获取"用户作答"列的索引
        header_row = 1  # Excel中的第一行
        user_answer_col_idx = None
        for idx, cell in enumerate(ws[header_row], 1):
            if cell.value == '用户作答':
                user_answer_col_idx = idx
                break

        if user_answer_col_idx is None:
            print("无法找到'用户作答'列")
            exit(1)

        # 创建临时文件列表，用于跟踪
        temp_files = []

        # 定义处理单个Excel行的函数
        def process_excel_row(args):
            row_idx, row, user_answer_col_idx = args
            pic_path = row[user_answer_col_idx-1].value
            temp_path = None
            
            if pic_path and os.path.exists(pic_path):
                try:
                    # 调整图片大小以适应Excel单元格
                    img = Image.open(pic_path)
                    max_height = 100  # 设置最大高度
                    width, height = img.size
                    new_width = int(width * (max_height / height)) if height > 0 else width
                    img = img.resize((new_width, max_height))
                    
                    # 创建临时文件以保存调整大小后的图片
                    temp_path = f"temp_{int(time.time())}_{threading.get_ident()}_{os.path.basename(pic_path)}"
                    img.save(temp_path)
                    
                    return row_idx, temp_path
                except Exception as e:
                    print(f"处理图片时出错 {pic_path}: {str(e)}")
            
            return row_idx, None

        # 使用线程池处理图片调整
        print("开始多线程处理图片大小调整...")
        row_data = [(row_idx, row, user_answer_col_idx) for row_idx, row in enumerate(ws.iter_rows(min_row=2), 2)]
        
        with concurrent.futures.ThreadPoolExecutor(max_workers=10) as executor:
            # 提交任务并获取结果
            results = list(tqdm(
                executor.map(process_excel_row, row_data),
                total=len(row_data),
                desc="调整图片大小"
            ))
        
        # 处理结果，添加图片到Excel
        for row_idx, temp_path in results:
            if temp_path:
                try:
                    temp_files.append(temp_path)
                    
                    # 将图片添加到Excel - 直接添加到"用户作答"列
                    img = XLImage(temp_path)
                    
                    # 计算图片位置 - 使用"用户作答"列
                    user_answer_col_letter = chr(64 + user_answer_col_idx)
                    cell_address = f"{user_answer_col_letter}{row_idx}"
                    
                    # 获取原始图片路径
                    pic_path = ws.cell(row=row_idx, column=user_answer_col_idx).value
                    
                    # 提取图片文件名作为单元格值
                    if pic_path:
                        file_name = os.path.basename(pic_path)
                        # 设置单元格值为图片文件名
                        ws[cell_address] = file_name
                    else:
                        ws[cell_address] = "图片"
                    
                    # 添加图片到工作表
                    ws.add_image(img, cell_address)
                    
                    # 获取最后添加的图片
                    img_idx = len(ws._images) - 1
                    if img_idx >= 0:
                        # 设置图片锚点，使其放在单元格内
                        img = ws._images[img_idx]
                        img.anchor = f"{user_answer_col_letter}{row_idx}"
                        # 设置图片不随单元格调整大小
                        img.anchor_type = 'oneCell'
                    
                except Exception as e:
                    print(f"添加图片到Excel时出错 {temp_path}: {str(e)}")

        # 调整列宽以更好地显示图片
        user_answer_col_letter = chr(64 + user_answer_col_idx)
        ws.column_dimensions[user_answer_col_letter].width = 20
        
        # 调整行高
        for i in range(2, ws.max_row + 1):
            ws.row_dimensions[i].height = 120  # 增加行高，确保图片显示完整

        # 保存带有图片的Excel - 处理权限问题
        try:
            embedded_file = f'{output_base}_核对.xlsx'
            # 检查文件是否已存在并且是否可以访问
            if os.path.exists(embedded_file):
                try:
                    # 尝试删除已存在的文件
                    os.remove(embedded_file)
                    print(f"已删除现有文件: {embedded_file}")
                except PermissionError:
                    # 如果无法删除，使用一个新的文件名
                    embedded_file = f'{output_base}_核对_{int(time.time())}.xlsx'
                    print(f"无法访问原文件，将使用新文件名: {embedded_file}")
            
            wb.save(embedded_file)
            print(f"处理完成！带有嵌入图片的结果已保存到 {embedded_file}")
        except Exception as e:
            print(f"保存文件时出错: {str(e)}")
            # 尝试使用备用文件名
            backup_file = f'{output_base}_embedded_{int(time.time())}.xlsx'
            wb.save(backup_file)
            print(f"已保存到备用文件: {backup_file}")
        
        # 最后清理所有临时文件
        print("清理临时文件...")
        for temp_file in temp_files:
            if os.path.exists(temp_file):
                try:
                    os.remove(temp_file)
                except Exception as e:
                    print(f"无法删除临时文件 {temp_file}: {str(e)}")
            
    except Exception as e:
        print(f"嵌入图片过程中发生错误: {str(e)}") 